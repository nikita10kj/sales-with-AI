from urllib import request
from django.contrib.auth.decorators import login_required
from django.views.generic import FormView, View,TemplateView
from django.contrib.auth import login, authenticate
from django.contrib import messages
from .models import EmailAttachment, EmailOTP, CustomUser, ProductService, ActivityLog,Signature
from .forms import EmailForm, OTPForm,SupportForm
from .utils import sendOTP, add_single_sender
from django.contrib.auth.views import PasswordResetView
from django.utils.timezone import now
from datetime import timedelta
from django.contrib.auth.mixins import LoginRequiredMixin,UserPassesTestMixin
from django.views.generic.edit import UpdateView
from django.urls import reverse_lazy
import os, json,re
from django.conf import settings
from django.shortcuts import redirect, get_object_or_404
from django.views.decorators.csrf import requires_csrf_token
from django.http import HttpResponseForbidden
from django.core.validators import URLValidator, EmailValidator
from django.core.exceptions import ValidationError
from collections import defaultdict
from generate_email.models import SentEmail,TargetAudience
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Count, Q, F, FloatField, ExpressionWrapper
from django.db.models.functions import Cast
from django.core.mail import EmailMessage
from generate_email.models import EmailSubscription
from generate_email.utils import create_subscription
from allauth.socialaccount.models import SocialAccount
from django.views import View
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse

class BlockDirectAccessMixin:
    def dispatch(self, request, *args, **kwargs):
        if request.method == "GET" and not request.META.get("HTTP_REFERER"):
            return redirect('/')
        return super().dispatch(request, *args, **kwargs)

@requires_csrf_token
def csrf_failure(request, reason=""):
    # Redirect to login page with error message
    messages.error(request, "Invalid session. Please log in again.")
    return redirect('login')

class LandingPageView(TemplateView):
    template_name = 'users/landingpage.html'
    title = "SellSharp"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = self.title
        return context


class HomeView(LoginRequiredMixin, TemplateView):
    template_name = 'users/dashboard.html'
    title = "Home"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = self.title
        user = self.request.user
        all_users = None

        # --- Subscription validation ---
        if EmailSubscription.objects.filter(user=user).exists():
            for sub in EmailSubscription.objects.filter(user=user):
                if sub.expires_at and sub.expires_at <= timezone.now():
                    create_subscription(user)
        else:
            create_subscription(user)

        # --- Email sending limit logic ---
        organization_domain = "jmsadvisory"
        user_email = (user.email or "").lower()
        is_jms_user = "@jmsadvisory" in user_email
        email_limit = None
        remaining_emails = None

        # Only apply the 500-email limit to external users
        if organization_domain not in user_email:
            total_sent_user = SentEmail.objects.filter(user=user).count()
            email_limit = 50
            remaining_emails = max(email_limit - total_sent_user, 0)
        else:
            total_sent_user = SentEmail.objects.filter(user=user).count()

        # --- Account filter from query param ---
        selected_account_id = self.request.GET.get("account")
        selected_account = None
        if selected_account_id:
            try:
                from allauth.socialaccount.models import SocialAccount as SA
                selected_account = SA.objects.get(pk=selected_account_id, user=user)
            except Exception:
                selected_account = None

        # Total sent emails
        if user.is_superuser:
            sent_emails = SentEmail.objects.all()
            campaign_types = TargetAudience.objects.all().exclude(framework__isnull=True).exclude(
                framework__exact='')
            recent_activities = ActivityLog.objects.all().order_by('-timestamp')[:5]

            today = now().date()

            all_users = CustomUser.objects.annotate(
                total_sent=Count('sent_email'),
                opened_count=Count('sent_email', filter=Q(sent_email__opened=True)),
                today_sent=Count(
                    'sent_email',
                    filter=Q(sent_email__created__date=today)
                ),
                today_opened=Count(
                    'sent_email',
                    filter=Q(sent_email__created__date=today, sent_email__opened=True)
                ),
            )
            

        else:
            sent_emails = SentEmail.objects.filter(user=user)
            # Filter by selected account if provided
            if selected_account:
                sent_emails = sent_emails.filter(sending_account=selected_account)
            campaign_types = TargetAudience.objects.filter(user=user).exclude(framework__isnull=True).exclude(
                framework__exact='')
            recent_activities = ActivityLog.objects.filter(user=user).order_by('-timestamp')[:5]

        # Calculate open rate
        total_sent=sent_emails.count()
        total_opened_emails = sent_emails.filter(opened=True).count()

        read_emails = total_opened_emails
        unread_emails = total_sent - total_opened_emails

        # open_rate = (total_opened_emails / total_sent * 100) if total_sent > 0 else 0
        # if user.email and user.email.lower() == "hardik@jmsadvisory.in":
        #     open_rate = 62.8

        # read_percentage = open_rate
        # unread_percentage = 100 - open_rate if total_sent > 0 else 0

        DEMO_EMAIL = "hardik@jmsadvisory.in"
        DEMO_OPEN_RATE = 62.8

        user_email = user.email.lower() if user.email else ""

        selected_email = ""
        if selected_account:
            if selected_account.provider == "google":
                selected_email = selected_account.extra_data.get("email", "").lower()
            else:
                selected_email = (
                    selected_account.extra_data.get("mail") or
                    selected_account.extra_data.get("userPrincipalName", "")
                ).lower()

        active_email = selected_email if selected_account else user_email

        if active_email == DEMO_EMAIL:
            open_rate = DEMO_OPEN_RATE
            read_percentage = DEMO_OPEN_RATE
            unread_percentage = round(100 - DEMO_OPEN_RATE, 1)
        else:
            open_rate = (total_opened_emails / total_sent * 100) if total_sent > 0 else 0
            read_percentage = open_rate
            unread_percentage = (100 - open_rate) if total_sent > 0 else 0

        # Get data for the last 7 days
        today = timezone.now().date()
        this_month_start = today.replace(day=1)
        yesterday = today - timedelta(days=1)

        today_sent = sent_emails.filter(
            created__date=today
        ).count()
        yesterday_sent = sent_emails.filter(created__date=yesterday).count()

        this_week_start = today - timedelta(days=today.weekday())  # Monday
        this_week_end = this_week_start + timedelta(days=6)

        last_week_start = this_week_start - timedelta(days=7)
        last_week_end = this_week_start - timedelta(days=1)

        last_month_end = this_month_start - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)

        # Counts
        this_week_sent = sent_emails.filter(
            created__date__gte=this_week_start,
            created__date__lte=this_week_end
        ).count()

        last_week_sent = sent_emails.filter(
            created__date__gte=last_week_start,
            created__date__lte=last_week_end
        ).count()

        this_month_sent = sent_emails.filter(
            created__date__gte=this_month_start
        ).count()

        last_month_sent = sent_emails.filter(
            created__date__gte=last_month_start,
            created__date__lte=last_month_end
        ).count()


        date_list = [today - timedelta(days=i) for i in range(6, -1, -1)]  # Last 7 days including today

        sent_data = []
        opened_data = []
        labels = []

        for date in date_list:
            start_date = datetime.combine(date, datetime.min.time())
            end_date = datetime.combine(date, datetime.max.time())
            
            sent_count = sent_emails.filter(
                created__gte=start_date,
                created__lte=end_date
            ).count()
            
            opened_count_day = sent_emails.filter(
                opened=True,
                created__gte=start_date,
                created__lte=end_date
            ).count()
            
            # sent_data.append(sent_count)
            sent_data.append(
                sent_emails.filter(created__range=(start_date, end_date)).count()
    )
            # opened_data.append(opened_count_day)
            opened_data.append(
                sent_emails.filter(
                opened=True,
                created__range=(start_date, end_date)
        ).count()
            )

            today_opened = sent_emails.filter(
            opened=True,
            opened_at__date=today
        ).count()
            labels.append(date.strftime("%b %d"))

            today_open_rate = (today_opened / today_sent * 100) if today_sent > 0 else 0

        # Get campaign types data - filtered by account via sent_email
        if selected_account:
            # Only TargetAudience that have sent emails via selected account
            filtered_ta_ids = sent_emails.values_list('target_audience_id', flat=True)
            campaign_types_filtered = campaign_types.filter(id__in=filtered_ta_ids)
        else:
            campaign_types_filtered = campaign_types

        campaign_data = campaign_types_filtered.values('framework').annotate(count=Count('framework')).order_by('-count')

        # Prepare data for chart
        campaign_labels = []
        campaign_counts = []
        
        for item in campaign_data:
            campaign_labels.append(item['framework'])
            campaign_counts.append(item['count'])

        # Get top performing campaigns (by open rate) - account filtered
        if selected_account:
            top_campaigns = campaign_types_filtered \
                .annotate(
                    total_sent=Count('sent_email', filter=Q(sent_email__sending_account=selected_account)),
                    opened_count=Count('sent_email', filter=Q(sent_email__opened=True, sent_email__sending_account=selected_account)),
                    open_rate=ExpressionWrapper(
                        Cast(F('opened_count'), FloatField()) / Cast(F('total_sent'), FloatField()) * 100,
                        output_field=FloatField()
                    )
                ) \
                .filter(total_sent__gt=0) \
                .order_by('-open_rate')[:4]
        else:
            top_campaigns = campaign_types \
                .annotate(
                    total_sent=Count('sent_email'),
                    opened_count=Count('sent_email', filter=Q(sent_email__opened=True)),
                    open_rate=ExpressionWrapper(
                        Cast(F('opened_count'), FloatField()) / Cast(F('total_sent'), FloatField()) * 100,
                        output_field=FloatField()
                    )
                ) \
                .filter(total_sent__gt=0) \
                .order_by('-open_rate')[:4]

        # Filter recent_activities by account if selected
        if selected_account and not user.is_superuser:
            recent_activities = ActivityLog.objects.filter(
                user=user,
                sent_email__sending_account=selected_account
            ).order_by('-timestamp')[:10]
        else:
            recent_activities = ActivityLog.objects.filter(
                user=user
            ).order_by('-timestamp')[:10]

        # Prepare top campaigns data
        top_campaigns_data = []
        for idx, campaign in enumerate(top_campaigns, start=1):
            top_campaigns_data.append({
                'rank': idx,
                'framework': campaign.framework,
                'service': campaign.selected_service or "General",
                'open_rate': round(campaign.open_rate, 1)
            })


        context.update({
            "total_sent": total_sent,
            "today_sent": today_sent,
            "yesterday_sent": yesterday_sent,
            "this_week_sent": this_week_sent,
            "last_week_sent": last_week_sent,
            "this_month_sent": this_month_sent,
            "last_month_sent": last_month_sent,
            "open_rate": round(open_rate, 1),
            "today_opened": today_opened,
            "today_open_rate": round(today_open_rate, 1),
            "read_emails": read_emails,
            "unread_emails": unread_emails,
            "read_percentage": round(read_percentage, 1),
            "unread_percentage": round(unread_percentage, 1),
            "chart_labels": labels,
            "sent_data": sent_data,
            "opened_data": opened_data,
            "campaign_labels": campaign_labels,
            "campaign_counts": campaign_counts,
            'top_campaigns': top_campaigns_data,
            'recent_activities': recent_activities,
            'all_users': all_users,
            'sent_emails': sent_emails,
            'email_limit': email_limit,
            'remaining_emails': remaining_emails,
            'total_sent_user': total_sent_user,
            'is_jms_user': is_jms_user,
            'selected_account_id': int(selected_account_id) if selected_account else None,
        })

        return context


class RegisterView(FormView):
    template_name = 'users/register.html'
    form_class = EmailForm
    title = "Sign Up"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = self.title

        return context

    def form_valid(self, form):
        email = form.cleaned_data['email'].lower()
        if CustomUser.objects.filter(email=email).exists():
            form.add_error('email', 'Email already registered. Try login.')
            return self.form_invalid(form)

        sendOTP(email)

        self.request.session['otp_email'] = email
        messages.success(self.request, 'OTP has been sent to your email.')
        return redirect('verify-otp')


class LoginView(FormView):
    template_name = 'users/login.html'
    form_class = EmailForm
    title = "Sign In"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = self.title

        return context

    def form_valid(self, form):
        email = form.cleaned_data['email'].lower()
        if not CustomUser.objects.filter(email=email).exists():
            form.add_error('email', 'Email not registered. Try registering.')
            return self.form_invalid(form)

        sendOTP(email)
        messages.success(self.request, "OTP has been sent to your email.")

        self.request.session['otp_email'] = email

        return redirect('verify-otp')

class VerifyOTPView(FormView):
    # verify with OTP or Password Login for Admin & Manager
    template_name = 'users/otp.html'
    form_class = OTPForm
    title = "Sign In"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = self.title

        return context

    def get_initial(self):
        return {'email': self.request.session.get('otp_email')}

    def form_valid(self, form):
        email = form.cleaned_data['email']
        otp_input = form.cleaned_data['otp']

        try:
            latest_otp = EmailOTP.objects.filter(email=email).latest('created_at')
        except EmailOTP.DoesNotExist:
            form.add_error(None, "Invalid OTP.")
            return self.form_invalid(form)

        if latest_otp.otp == otp_input and latest_otp.is_valid():

            user, created = CustomUser.objects.get_or_create(email=email)
            if created:
                redirect_url = reverse_lazy('user-details', kwargs={'pk': user.id})
            else:
                if user.full_name:
                    redirect_url = 'home'
                else:
                    redirect_url = reverse_lazy('user-details', kwargs={'pk': user.id})

            user.backend = 'django.contrib.auth.backends.ModelBackend'  # or your actual backend path

            login(self.request, user)
            EmailOTP.objects.filter(email=email).delete()

            return redirect(redirect_url)  # Or your home

        form.add_error('otp', 'Invalid or expired OTP')
        return self.form_invalid(form)

    def form_invalid(self, form):
        return super().form_invalid(form)


class ResendOTPView(View):
    def post(self, request, *args, **kwargs):
        email = request.POST.get('email')
        if not email or email == 'None' or email == '':
            messages.error(request, "Email is missing.")
            return redirect('login')  # or appropriate page

        # Rate limiting: no more than 3 OTPs per 5 minutes
        recent_otps = EmailOTP.objects.filter(email=email, created_at__gte=now() - timedelta(minutes=5))
        if recent_otps.count() >= 3:
            messages.error(request, "Too many OTP requests. Try again later.")
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # Create new OTP
        sendOTP(email)

        messages.success(request, "A new OTP has been sent to your email.")
        return redirect('verify-otp')

from django.core.validators import URLValidator
from django.core.exceptions import ValidationError
class UserDetailsView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = CustomUser
    fields = [
        'full_name',
        'company_name',
        'company_url',
        'company_linkedin_url',
        'user_linkedin_url',
        'contact'
    ]
    template_name = 'users/user_details.html'
    success_url = reverse_lazy('home')
    title = "User Details"

    # Helper function to prepend https:// if missing from the URL
    def normalize_url(self, url):
        if url and not url.startswith(('http://', 'https://')):
            return f'https://{url.strip()}'
        return url.strip() if url else ''

    # Override get_form to sanitize URL fields before form validation
    def get_form(self, form_class=None):
        print("→ get_form called!")
        form = super().get_form(form_class)

        # Only modify data when POST request
        if self.request.method == "POST":
            data = self.request.POST.copy()
            validator = URLValidator()

            for field in ['company_url', 'company_linkedin_url', 'user_linkedin_url']:
                original = data.get(field)
                normalized = self.normalize_url(original)

                try:
                    if normalized:
                        validator(normalized)
                except ValidationError:
                    form.add_error(field, "Enter a valid URL.")

                data[field] = normalized

            # recreate form with updated data
            form = form.__class__(
                data,
                instance=self.get_object(),
                files=self.request.FILES
            )

        return form

    # Restrict access: users can only edit their own profile
    def test_func(self):
        """
        Restrict access to only the owner of the profile.
        """
        user = get_object_or_404(CustomUser, pk=self.kwargs['pk'])
        return self.request.user == user

    def handle_no_permission(self):
        """
        Custom response for unauthorized access.
        """
        if not self.request.user.is_authenticated:
            return super().handle_no_permission()
        return HttpResponseForbidden("You do not have permission to access this page.")

    def get_object(self, queryset=None):
        """
        Always use the logged-in user as the object.
        """
        return self.request.user

    def get_context_data(self, **kwargs):
        """
        Add additional context like title and services to the template.
        """
        context = super().get_context_data(**kwargs)
        context["title"] = self.title
        context["services"] = ProductService.objects.filter(user=self.request.user)
        return context

    def form_valid(self, form):
        """
        Handles successful form submission, updates services, and shows a success message.
        """
        response = super().form_valid(form)
        user = self.request.user

        # Clear existing services
        ProductService.objects.filter(user=user).delete()

        # Collect new services using regex and defaultdict
        services_data = defaultdict(dict)
        service_pattern = re.compile(r'service_name_(\d+)')

        for key, value in self.request.POST.items():
            match = service_pattern.match(key)
            if match:
                idx = match.group(1)
                services_data[idx]['service_name'] = value.strip()
                services_data[idx]['product_url'] = self.normalize_url(self.request.POST.get(f'product_url_{idx}', ''))
                services_data[idx]['product_usp'] = self.request.POST.get(f'product_usp_{idx}', '').strip()

        # Save valid services
        # for service in services_data.values():
        #     if service.get('service_name') and service.get('product_url'):
        #         ProductService.objects.create(
        #             user=user,
        #             service_name=service['service_name'],
        #             product_url=service['product_url'],
        #             product_usp=service.get('product_usp', '')
        #         )

        validator = URLValidator()

        for service in services_data.values():
            if service.get('service_name') and service.get('product_url'):
                try:
                    validator(service['product_url'])
                except ValidationError:
                    messages.error(
                        self.request,
                        f"Invalid URL for service: {service.get('service_name')}"
                    )
                    return self.form_invalid(form)

                ProductService.objects.create(
                    user=user,
                    service_name=service['service_name'],
                    product_url=service['product_url'],
                    product_usp=service.get('product_usp', '')
                )

        # Optional custom action and user feedback
        add_single_sender(user)
        messages.success(self.request, "Your details and services have been updated successfully.")

        return response

class ProfileView(BlockDirectAccessMixin,LoginRequiredMixin, View):

    def normalize_url(self, url):
        if url and not url.startswith(('http://', 'https://')):
            return f'https://{url.strip()}'
        return url.strip() if url else ''

    def get(self, request):
        google_accounts = SocialAccount.objects.filter(
            user=request.user,
            provider="google"
        )

        microsoft_accounts = SocialAccount.objects.filter(
        user=request.user,
        provider="microsoft"
        )


        services = ProductService.objects.filter(
            user=request.user
        ).order_by('id')

        signatures = Signature.objects.filter(
        user=request.user
        ).order_by('id')
        
        attachments = EmailAttachment.objects.filter(user=request.user)

        return render(request, 'users/profile.html', {
            'user': request.user,
            'services': services,
            "signatures": signatures,
            "google_accounts": google_accounts,
            "microsoft_accounts": microsoft_accounts,
            'attachments': attachments,
        })

    
    def post(self, request):
        user = request.user
        if 'personal_submit' in request.POST:

            errors = {}
            changes = {}

            # Email
            email = request.POST.get('email', '').strip()
            try:
                EmailValidator()(email)
                if user.email != email:
                    if CustomUser.objects.exclude(id=user.id).filter(email=email).exists():
                        errors['email'] = "Email already exists"
                    changes['email'] = {'old': user.email, 'new': email}
                user.email = email
            except ValidationError:
                errors['email'] = 'Invalid email format'

            # Full Name
            full_name = request.POST.get('full_name', '').strip()
            if user.full_name != full_name:
                changes['full_name'] = {'old': user.full_name, 'new': full_name}
            user.full_name = full_name

            # Contact
            contact = request.POST.get('contact', '').strip()
            if user.contact != contact:
                changes['contact'] = {'old': user.contact, 'new': contact}
            user.contact = contact

            # Company Name
            company_name = request.POST.get('company_name', '').strip()
            if user.company_name != company_name:
                changes['company_name'] = {'old': user.company_name, 'new': company_name}
            user.company_name = company_name

            # Company URL
            company_url = self.normalize_url(request.POST.get('company_url', ''))
            if user.company_url != company_url:
                changes['company_url'] = {'old': user.company_url, 'new': company_url}
            user.company_url = company_url

            # Company LinkedIn
            company_linkedin_url = self.normalize_url(
                request.POST.get('company_linkedin_url', '')
            )
            if user.company_linkedin_url != company_linkedin_url:
                changes['company_linkedin_url'] = {
                    'old': user.company_linkedin_url,
                    'new': company_linkedin_url
                }
            user.company_linkedin_url = company_linkedin_url

            # User LinkedIn
            user_linkedin_url = self.normalize_url(
                request.POST.get('user_linkedin_url', '')
            )
            if user.user_linkedin_url != user_linkedin_url:
                changes['user_linkedin_url'] = {
                    'old': user.user_linkedin_url,
                    'new': user_linkedin_url
                }
            user.user_linkedin_url = user_linkedin_url

            # Errors → return page
            if errors:
                services = ProductService.objects.filter(user=user)
                return render(request, 'users/profile.html', {
                    'user': user,
                    'services': services,
                    'errors': errors
                })

            user.save()

            if changes:
                ActivityLog.objects.get_or_create(
                    user=user,
                    action="PROFILE_UPDATED",
                    description="Profile updated"
                )

            messages.success(request, "Personal information updated successfully!")
            return redirect('profile')

        
        elif 'knowledge_submit' in request.POST:
            ProductService.objects.filter(user=user).delete()
            services_data = defaultdict(dict)
            pattern = re.compile(r'service_name_(\d+)')

            for key, value in request.POST.items():
                match = pattern.match(key)
                if match:
                    idx = match.group(1)
                    services_data[idx]['service_name'] = value.strip()
                    services_data[idx]['product_url'] = self.normalize_url(
                        request.POST.get(f'product_url_{idx}', '')
                    )
                    services_data[idx]['product_usp'] = request.POST.get(
                        f'product_usp_{idx}', ''
                    ).strip()

            for service in services_data.values():
                if service.get('service_name') and service.get('product_url'):
                    ProductService.objects.create(
                        user=user,
                        service_name=service['service_name'],
                        product_url=service['product_url'],
                        product_usp=service.get('product_usp', '')
                    )

            messages.success(request, "Knowledge Base saved successfully!")
            return redirect('profile')
        
        # elif 'signature_submit' in request.POST:
        #     old_sigs = list(Signature.objects.filter(user=user).order_by('id'))
        #     old_photos = {i: s.photo for i, s in enumerate(old_sigs)}

        #     Signature.objects.filter(user=user).delete()

        #     for key, value in request.POST.items():
        #         if key.startswith('signature_') and not key.startswith('signature_name_') and value.strip():
        #             idx = int(key.split('_')[1])

        #             name = request.POST.get(f'signature_name_{idx}', '').strip()
        #             photo = request.FILES.get(f'signature_photo_{idx}')

        #             if not photo:
        #                 photo = old_photos.get(idx)

        #             Signature.objects.create(
        #                 user=user,
        #                 name=name,
        #                 signature=value.strip(),
        #                 photo=photo
        #             )

            # messages.success(request, "Signatures saved successfully!")
            # return redirect('profile')

        elif 'signature_submit' in request.POST:
            old_sigs = list(Signature.objects.filter(user=user).order_by('id'))
            old_photos = {i: s.photo for i, s in enumerate(old_sigs)}

            has_signature = False

            # 🔍 Check if at least one signature exists
            for key, value in request.POST.items():
                if key.startswith('signature_') and not key.startswith('signature_name_') and value.strip():
                    has_signature = True
                    break

            # ❌ No signature case
            if not has_signature:
                messages.error(request, "No signature added")
                return redirect('profile')

            # ✅ If signature exists → then delete & save
            Signature.objects.filter(user=user).delete()

            for key, value in request.POST.items():
                if key.startswith('signature_') and not key.startswith('signature_name_') and value.strip():
                    idx = int(key.split('_')[1])

                    name = request.POST.get(f'signature_name_{idx}', '').strip()
                    photo = request.FILES.get(f'signature_photo_{idx}')

                    if not photo:
                        photo = old_photos.get(idx)

                    Signature.objects.create(
                        user=user,
                        name=name,
                        signature=value.strip(),
                        photo=photo
                    )

            messages.success(request, "Signatures saved successfully!")
            return redirect('profile')

        
        elif 'attachment_submit' in request.POST:
            files = request.FILES.getlist('attachments')

            for f in files:
                EmailAttachment.objects.create(
                    user=user,
                    file=f,
                    original_name=f.name
                )

            messages.success(request, "Attachments uploaded successfully!")
            return redirect('profile')


        return redirect('profile')
        
class PrivacyPolicyView(View):
    template_name='users/privacypolicy.html'

    def get(self, request):
        return render(request, self.template_name)

class LearningHubView(View):
    template_name='users/learninghub.html'

    def get(self, request):
        return render(request, self.template_name)
class TermsConditionsView(View):
    template_name='users/termsconditions.html'

    def get(self, request):
        return render(request, self.template_name)
class RefundPolicyView(View):
    template_name='users/refundpolicy.html'

    def get(self, request):
        return render(request, self.template_name)

class ContactUsView(View):
    template_name='users/contact_us.html'

    def get(self, request):
        return render(request, self.template_name)
    
    def post(self, request):
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        phone = request.POST.get('phone', '')
        subject = request.POST.get('subject', '')
        message = request.POST.get('message', '')
        
        if name and email and subject and message:
            try:
                # Send email notification to admin
                admin_email = 'harshil@jmsadvisory.in'
                email_body = f"""
                New Contact Us Message from SellSharp Website:
                
                Name: {name}
                Email: {email}
                Phone: {phone if phone else 'Not provided'}
                Subject: {subject}
                
                Message:
                {message}
                """
                
                email_msg = EmailMessage(
                    subject=f"New Contact Form Submission: {subject}",
                    body=email_body,
                    from_email='noreply@sellsharp.co',
                    to=[admin_email],
                    reply_to=[email]
                )
                email_msg.send(fail_silently=True)
                
                messages.success(request, 'Thank you! Your message has been sent successfully. We will get back to you soon.')
            except Exception as e:
                messages.warning(request, 'Your message was received, but we had an issue sending it. Please try emailing us directly at info@jmsadvisory.in')
        else:
            messages.error(request, 'Please fill in all required fields.')
        
        return render(request, self.template_name)

class PricingView(BlockDirectAccessMixin,View):
    template_name = "users/pricing.html"

    def get(self, request):
        plans = [
            {
                "name": "Starter",
                "tagline": "For individuals getting started",
                "price_month": 0,
                "price_year": 0,
                "cta": "Start Free Trial",
                "popular": False,
                "features": [
                    "AI Email Generator",
                    "Basic personalization",
                    "Limited templates",
                    "Community support",
                ],
            },
            {
                "name": "Pro",
                "tagline": "Best for sales reps & freelancers",
                "price_month": 299,
                "price_year": 2990,
                "cta": "Start Free Trial",
                "popular": True,
                "features": [
                    "Everything in Starter",
                    "Smart personalization",
                    "Tone & style control",
                    "Performance analytics",
                ],
            },
            {
                "name": "Team",
                "tagline": "For growing teams & agencies",
                "price_month": 799,
                "price_year": 7990,
                "cta": "Talk to Sales",
                "popular": False,
                "features": [
                    "Everything in Pro",
                    "Shared templates & sequences",
                    "Team management",
                    "Priority support",
                ],
            },
        ]

        context = {"plans": plans}
        return render(request, self.template_name, context)
    
import json, hmac, hashlib
from django.conf import settings
from django.http import JsonResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
import razorpay
from .models import RazorpayCreditOrder, UserWallet

razorpay_client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))
MIN_CREDITS = 1000

@login_required
@require_POST
def razorpay_create_order(request):
    data = json.loads(request.body or "{}")
    credits = int(data.get("credits", 0))
    if credits < MIN_CREDITS:
        return JsonResponse({"error": "Minimum purchase is 1000 credits"}, status=400)

    amount_rupees = credits
    amount_paise = amount_rupees * 100

    rp_order = razorpay_client.order.create({
        "amount": amount_paise,
        "currency": "INR",
        "receipt": f"credits_{request.user.id}_{credits}",
        "payment_capture": 1,
        "notes": {"user_id": str(request.user.id), "credits": str(credits)}
    })

    RazorpayCreditOrder.objects.create(
        user=request.user,
        credits=credits,
        amount_rupees=amount_rupees,
        amount_paise=amount_paise,
        razorpay_order_id=rp_order["id"],
        status="CREATED",
    )

    return JsonResponse({
        "key_id": settings.RAZORPAY_KEY_ID,
        "order_id": rp_order["id"],
        "amount": amount_paise,
        "currency": "INR",
        "credits": credits,
        "name": "SellSharp",
        "description": f"Buy {credits} credits",
        "prefill": {"name": request.user.get_full_name(), "email": request.user.email},
    })


@login_required
@require_POST
def razorpay_verify_payment(request):
    data = json.loads(request.body or "{}")
    oid = data.get("razorpay_order_id")
    pid = data.get("razorpay_payment_id")
    sig = data.get("razorpay_signature")

    if not (oid and pid and sig):
        return JsonResponse({"error": "Missing Razorpay fields"}, status=400)

    try:
        order = RazorpayCreditOrder.objects.select_for_update().get(user=request.user, razorpay_order_id=oid)
    except RazorpayCreditOrder.DoesNotExist:
        return JsonResponse({"error": "Order not found"}, status=404)

    if order.status == "PAID":
        wallet = UserWallet.objects.get(user=request.user)
        return JsonResponse({"status": "success", "credits_added": 0, "new_balance": wallet.credits})

    msg = f"{oid}|{pid}".encode("utf-8")
    expected = hmac.new(settings.RAZORPAY_KEY_SECRET.encode("utf-8"), msg, hashlib.sha256).hexdigest()

    if expected != sig:
        order.status = "FAILED"
        order.save(update_fields=["status"])
        return JsonResponse({"error": "Signature mismatch"}, status=400)

    wallet, _ = UserWallet.objects.get_or_create(user=request.user, defaults={"credits": 500})
    wallet.credits += order.credits
    wallet.save(update_fields=["credits", "updated_at"])

    order.razorpay_payment_id = pid
    order.razorpay_signature = sig
    order.status = "PAID"
    order.paid_at = timezone.now()
    order.save(update_fields=["razorpay_payment_id", "razorpay_signature", "status", "paid_at"])

    return JsonResponse({"status": "success", "credits_added": order.credits, "new_balance": wallet.credits})


class SupportView(BlockDirectAccessMixin,LoginRequiredMixin, FormView):    
    template_name = "users/support.html"    
    form_class = SupportForm    
    def get_context_data(self, **kwargs):        
        context = super().get_context_data(**kwargs)        
        return context    
    
    def form_valid(self, form):        
        email = form.cleaned_data["email"]        
        subject = form.cleaned_data["subject"]        
        message = form.cleaned_data["message"]        
        email_msg = EmailMessage(            
            subject=f"[Support] {subject}",            
            body=message,            
            to=["resumate1nfo1@gmail.com"],            
            reply_to=[email],        
            )        
        email_msg.send(fail_silently=False)        
        messages.success(self.request, "We have received your message!")        
        return redirect('support')    
    
    def form_invalid(self, form):        
        messages.error(self.request, "There was a problem with your submission.")        
        return self.render_to_response(self.get_context_data(form=form))

@login_required
def thirdparty_redirect(request):
    """
    Override allauth 3rdparty page
    Always redirect to profile
    """
    return redirect('login')

@login_required
def dashboard(request):
    user = request.user

    # --- Connected email accounts (Google + Microsoft) ---
    connected_accounts = SocialAccount.objects.filter(
        user=user, provider__in=["google", "microsoft"]
    )

    # Selected account filter (from ?account=<id> query param)
    selected_account_id = request.GET.get("account")
    selected_account = None
    if selected_account_id:
        try:
            selected_account = connected_accounts.get(pk=selected_account_id)
        except SocialAccount.DoesNotExist:
            selected_account = None

    # Base queryset — filter by selected account if chosen
    qs = SentEmail.objects.filter(user=user)
    if selected_account:
        qs = qs.filter(sending_account=selected_account)

    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    last_week = today - timedelta(days=7)
    last_month = today - timedelta(days=30)

    total_sent       = qs.count()
    opened_emails    = qs.filter(opened=True).count()
    unread_emails    = qs.filter(opened=False).count()
    today_sent       = qs.filter(created__date=today).count()
    yesterday_sent   = qs.filter(created__date=yesterday).count()
    last_week_sent   = qs.filter(created__date__gte=last_week).count()
    last_month_sent  = qs.filter(created__date__gte=last_month).count()
    today_opened     = qs.filter(opened=True, opened_at__date=today).count()

    open_rate        = round((opened_emails / total_sent) * 100, 2) if total_sent else 0
    read_percentage  = open_rate
    unread_percentage = round(100 - open_rate, 2) if total_sent else 0
    today_open_rate  = round((today_opened / today_sent) * 100, 2) if today_sent else 0

    # Build a helper list for template: account id + display email
    account_list = []
    for acc in connected_accounts:
        if acc.provider == "google":
            display = acc.extra_data.get("email", "")
        else:
            display = acc.extra_data.get("mail") or acc.extra_data.get("userPrincipalName", "")
        account_list.append({"id": acc.pk, "email": display, "provider": acc.provider})

    context = {
        "total_sent": total_sent,
        "opened_emails": opened_emails,
        "unread_emails": unread_emails,
        "read_emails": opened_emails,
        "open_rate": open_rate,
        "today_sent": today_sent,
        "yesterday_sent": yesterday_sent,
        "last_week_sent": last_week_sent,
        "last_month_sent": last_month_sent,
        "today_opened": today_opened,
        "today_open_rate": today_open_rate,
        "read_percentage": read_percentage,
        "unread_percentage": unread_percentage,
        "connected_accounts": account_list,
        "selected_account_id": int(selected_account_id) if selected_account_id and selected_account else None,
    }

    return render(request, "users/dashboard.html", context)

@login_required
def remove_google_account(request, pk):
    acc = get_object_or_404(SocialAccount, pk=pk, user=request.user)
    acc.delete()
    return redirect("profile")   # tamaru profile page


from django.contrib.auth.mixins import UserPassesTestMixin
from django.views.generic import TemplateView

class SuperUserRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser


from django.db.models import Count
from django.utils import timezone
from datetime import timedelta
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from .models import CustomUser


class AdminDashboardView(BlockDirectAccessMixin,LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = "users/admin_dashboard.html"

    def test_func(self):
        return self.request.user.is_superuser

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        now = timezone.now()
        today = now.date()
        week_start = today - timedelta(days=7)
        month_start = today.replace(day=1)

        users = CustomUser.objects.filter(
            is_superuser=False
        ).annotate(
            email_count=Count("sent_email")
        ).order_by("-last_login")


        context["users"] = users
        context["superuser"] = CustomUser.objects.filter(is_superuser=True).first()
        context["today_logins"] = CustomUser.objects.filter(last_login__date=today).count()
        context["week_logins"] = CustomUser.objects.filter(last_login__date__gte=week_start).count()
        context["month_logins"] = CustomUser.objects.filter(last_login__date__gte=month_start).count()

        return context

@login_required
def delete_attachment(request, pk):
    attachment = get_object_or_404(
        EmailAttachment,
        pk=pk,
        user=request.user
    )

    # physical file delete
    if attachment.file:
        attachment.file.delete(save=False)

    attachment.delete()
    messages.success(request, "Attachment deleted successfully!")
    return redirect("profile")

@login_required
def list_user_attachments(request):
    attachments = EmailAttachment.objects.filter(user=request.user)

    data = [
        {
            "id": att.id,
            "name": att.original_name,
            "url": att.file.url
        }
        for att in attachments
    ]

    return JsonResponse({"attachments": data})

@login_required
def delete_signature(request, pk):
    if request.method == "POST":
        Signature.objects.filter(id=pk, user=request.user).delete()
        return JsonResponse({"success": True})
    return JsonResponse({"success": False})


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

class AdminDashboardExportView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Export admin dashboard data to Excel"""
    
    def test_func(self):
        return self.request.user.is_superuser
    
    def get(self, request):
        # Get all data same as AdminDashboardView
        now = timezone.now()
        today = now.date()
        week_start = today - timedelta(days=7)
        month_start = today.replace(day=1)
        
        users = CustomUser.objects.filter(
            is_superuser=False
        ).annotate(
            email_count=Count("sent_email")
        ).order_by("-last_login")
        
        superuser = CustomUser.objects.filter(is_superuser=True).first()
        today_logins = CustomUser.objects.filter(last_login__date=today).count()
        week_logins = CustomUser.objects.filter(last_login__date__gte=week_start).count()
        month_logins = CustomUser.objects.filter(last_login__date__gte=month_start).count()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard"
        
        # Define styles
        header_fill = PatternFill(start_color="6F6AE1", end_color="6F6AE1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        stat_fill = PatternFill(start_color="E8E8FF", end_color="E8E8FF", fill_type="solid")
        stat_font = Font(bold=True, size=11)
        label_font = Font(color="6B7280", size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = "👑 ADMIN DASHBOARD REPORT"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Generate timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=9, color="6B7280")
        
        # Superuser info
        row = 4
        ws[f'A{row}'] = "Super User"
        ws[f'A{row}'].font = stat_font
        ws[f'B{row}'] = superuser.email if superuser else "N/A"
        
        # Stats section
        row = 6
        ws[f'A{row}'] = "LOGIN STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws.merge_cells(f'A{row}:B{row}')
        
        row = 7
        stats_data = [
            ("Today Logins", today_logins),
            ("This Week Logins", week_logins),
            ("This Month Logins", month_logins),
        ]
        
        for label, value in stats_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = stat_font
            ws[f'B{row}'].fill = stat_fill
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            row += 1
        
        # Users table
        row = 12
        ws[f'A{row}'] = f"ALL USERS ({users.count()})"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws.merge_cells(f'A{row}:D{row}')
        
        # Table headers
        row = 13
        headers = ["User", "Email", "Emails Sent", "Last Login"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Table data
        row = 14
        for user in users:
            # User name
            user_name = user.get_full_name() or user.email
            ws[f'A{row}'] = user_name
            ws[f'A{row}'].border = border
            
            # Email
            ws[f'B{row}'] = user.email
            ws[f'B{row}'].border = border
            
            # Emails sent
            ws[f'C{row}'] = user.email_count
            ws[f'C{row}'].border = border
            ws[f'C{row}'].alignment = Alignment(horizontal='center')
            
            # Last login
            if user.last_login:
                last_login = user.last_login.strftime('%Y-%m-%d %H:%M')
            else:
                last_login = "Never"
            ws[f'D{row}'] = last_login
            ws[f'D{row}'].border = border
            ws[f'D{row}'].alignment = Alignment(horizontal='center')
            
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="admin_dashboard_{today.strftime("%Y%m%d")}.xlsx"'
        
        wb.save(response)
        return response